import hashlib
import os

from openpyxl import Workbook, load_workbook


excel_path = os.getenv("AUTH_FIXTURE_XLSX", ".\\authDB\\auth.xlsx")


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def add_student_to_excel(auth_dict):
    ws, wb = create_excel()
    for student_id, plain_password in auth_dict.items():
        if check_registered_user(student_id, ws):
            continue
        hashed_pw = hash_password(plain_password)
        ws.append([student_id, hashed_pw])
        wb.save(excel_path)
        print(f"Registered: {student_id}")


def make_sample_dictionary():
    return {
        "<student-id-1>": "<sample-password-1>",
        "<student-id-2>": "<sample-password-2>",
    }


def create_excel():
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "auth"
        ws.append(["student_id", "password_hash"])
    else:
        wb = load_workbook(excel_path)
        ws = wb.active
    return ws, wb


def check_registered_user(student_id, ws):
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == str(student_id):
            print(f"Already registered: {student_id}")
            return True
    return False


if __name__ == "__main__":
    add_student_to_excel(make_sample_dictionary())
